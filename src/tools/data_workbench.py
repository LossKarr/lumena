"""
data_workbench.py — Moteur de manipulation tabulaire pour Lumena.

V2.1 (lecture + profilage) :
  - Profile un fichier CSV / XLSX / JSON
  - Inférence dtype (int / float / date / text)
  - Détection colonnes dates / territoires / numériques
  - Auto-detect encoding (utf-8 → latin-1 → cp1252)
  - Auto-detect séparateur CSV (`,` `;` `\\t` `|`)
  - Lecture provenance data.gouv via sidecar `<file>.datagouv.json`

Zéro dépendance externe en V2.1 (csv + json stdlib + openpyxl déjà disponible).
Les opérations query/filter/aggregate/export arrivent en V2.2-V2.4.
"""

from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Optional


# ─── Configuration ──────────────────────────────────────────────────────

MAX_ROWS_PROFILE = 500_000  # override via env LUMENA_DATA_WB_MAX_ROWS
SAMPLE_SIZE = 5


def _max_rows_default() -> int:
    import os
    try:
        return int(os.getenv("LUMENA_DATA_WB_MAX_ROWS", str(MAX_ROWS_PROFILE)))
    except ValueError:
        return MAX_ROWS_PROFILE


# ─── Heuristiques détection ─────────────────────────────────────────────

_DATE_PATTERNS = [
    re.compile(r"^\d{4}-\d{2}-\d{2}"),       # ISO 2024-01-31
    re.compile(r"^\d{1,2}/\d{1,2}/\d{4}"),   # FR 31/01/2024
    re.compile(r"^\d{1,2}-\d{1,2}-\d{4}"),   # 31-01-2024
]
_INT_RE = re.compile(r"^-?\d+$")
_FLOAT_RE = re.compile(r"^-?\d+[.,]\d+$")

_TERRITORY_KEYWORDS = (
    "commune", "departement", "département", "region", "région",
    "insee", "code_postal", "siren", "siret", "iris", "epci",
    "code_commune", "code_departement", "code_region",
)

_DATE_NAME_KEYWORDS = ("date", "annee", "année", "year", "jour", "mois")


# ─── Dataclasses ────────────────────────────────────────────────────────

@dataclass
class ColumnProfile:
    name: str
    dtype: str  # int / float / date / text
    null_count: int
    null_pct: float
    sample: List[str]
    is_territory: bool
    is_date_probable: bool
    is_numeric_probable: bool


@dataclass
class ProfileResult:
    path: str
    format: str  # csv / xlsx / json
    rows: int
    cols: int
    columns: List[ColumnProfile]
    sample_rows: List[dict]
    truncated: bool
    encoding_used: Optional[str] = None
    separator_used: Optional[str] = None
    limits: List[str] = field(default_factory=list)
    provenance: Optional[dict] = None
    error: Optional[str] = None


# ─── Helpers détection ──────────────────────────────────────────────────

def _detect_encoding(path: Path) -> str:
    # Détecte un BOM UTF-8 explicitement → utf-8-sig (strip auto du BOM)
    try:
        with path.open("rb") as f:
            head = f.read(3)
        if head.startswith(b"\xef\xbb\xbf"):
            return "utf-8-sig"
    except OSError:
        pass
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            with path.open("r", encoding=enc) as f:
                f.read(8192)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"


def _strip_bom(s: str) -> str:
    """Retire un BOM résiduel en début de chaîne (sécurité supplémentaire)."""
    if s and s.startswith("﻿"):
        return s.lstrip("﻿")
    return s


def _clean_headers(headers: List[str]) -> List[str]:
    """Nettoie BOM + espaces périphériques des en-têtes de colonnes."""
    return [_strip_bom(h).strip() if isinstance(h, str) else h for h in headers]


def _detect_separator(path: Path, encoding: str) -> str:
    with path.open("r", encoding=encoding, errors="replace") as f:
        sample = f.read(8192)
    candidates = {sep: sample.count(sep) for sep in (",", ";", "\t", "|")}
    best = max(candidates, key=candidates.get)
    return best if candidates[best] > 0 else ","


def _infer_dtype(values: List[str]) -> str:
    non_null = [v for v in values if v not in ("", None)]
    if not non_null:
        return "text"
    int_count = sum(1 for v in non_null if _INT_RE.match(v.strip()))
    float_count = sum(1 for v in non_null if _FLOAT_RE.match(v.strip()))
    date_count = sum(
        1 for v in non_null if any(p.match(v.strip()) for p in _DATE_PATTERNS)
    )
    n = len(non_null)
    if int_count / n > 0.95:
        return "int"
    if (int_count + float_count) / n > 0.95:
        return "float"
    if date_count / n > 0.8:
        return "date"
    return "text"


def _is_territory_column(name: str) -> bool:
    n = name.lower()
    return any(kw in n for kw in _TERRITORY_KEYWORDS)


def _is_date_name(name: str) -> bool:
    n = name.lower()
    return any(kw in n for kw in _DATE_NAME_KEYWORDS)


def _profile_columns(headers: List[str], rows: List[List[str]]) -> List[ColumnProfile]:
    n_rows = len(rows)
    out: List[ColumnProfile] = []
    for i, name in enumerate(headers):
        values = [r[i] if i < len(r) else "" for r in rows]
        non_empty = [v for v in values if v not in ("", None)]
        nulls = n_rows - len(non_empty)
        dtype = _infer_dtype(values)
        # Echantillons : valeurs distinctes non vides, max 3
        sample = list(dict.fromkeys(non_empty))[:3]
        out.append(ColumnProfile(
            name=name,
            dtype=dtype,
            null_count=nulls,
            null_pct=round(100 * nulls / n_rows, 1) if n_rows else 0.0,
            sample=sample,
            is_territory=_is_territory_column(name),
            is_date_probable=dtype == "date" or _is_date_name(name),
            is_numeric_probable=dtype in ("int", "float"),
        ))
    return out


# ─── Readers par format ─────────────────────────────────────────────────

def _profile_csv(path: Path, max_rows: int) -> ProfileResult:
    encoding = _detect_encoding(path)
    separator = _detect_separator(path, encoding)
    limits: List[str] = []
    rows: List[List[str]] = []

    with path.open("r", encoding=encoding, errors="replace", newline="") as f:
        reader = csv.reader(f, delimiter=separator)
        try:
            headers = _clean_headers(next(reader))
        except StopIteration:
            return ProfileResult(
                path=str(path), format="csv", rows=0, cols=0,
                columns=[], sample_rows=[], truncated=False,
                encoding_used=encoding, separator_used=separator,
                error="fichier vide",
            )
        truncated = False
        for i, row in enumerate(reader):
            if i >= max_rows:
                truncated = True
                limits.append(f"Tronqué à {max_rows} lignes (fichier plus grand)")
                break
            rows.append(row)

    columns = _profile_columns(headers, rows)
    sample_rows = [dict(zip(headers, r)) for r in rows[:SAMPLE_SIZE]]
    return ProfileResult(
        path=str(path), format="csv", rows=len(rows), cols=len(headers),
        columns=columns, sample_rows=sample_rows, truncated=truncated,
        encoding_used=encoding, separator_used=separator, limits=limits,
    )


def _profile_xlsx(path: Path, max_rows: int) -> ProfileResult:
    try:
        import openpyxl
    except ImportError:
        return ProfileResult(
            path=str(path), format="xlsx", rows=0, cols=0,
            columns=[], sample_rows=[], truncated=False,
            error="openpyxl non installé",
        )
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb.active
    limits: List[str] = []
    rows_iter = ws.iter_rows(values_only=True)
    try:
        first = next(rows_iter)
    except StopIteration:
        wb.close()
        return ProfileResult(
            path=str(path), format="xlsx", rows=0, cols=0,
            columns=[], sample_rows=[], truncated=False,
            error="feuille vide",
        )
    headers = _clean_headers([str(c) if c is not None else "" for c in first])
    rows: List[List[str]] = []
    truncated = False
    for i, row in enumerate(rows_iter):
        if i >= max_rows:
            truncated = True
            limits.append(f"Tronqué à {max_rows} lignes")
            break
        rows.append([str(c) if c is not None else "" for c in row])

    if wb.sheetnames and len(wb.sheetnames) > 1:
        limits.append(
            f"Feuille profilée : '{ws.title}'. Autres feuilles ignorées : "
            f"{[s for s in wb.sheetnames if s != ws.title]}"
        )
    wb.close()

    columns = _profile_columns(headers, rows)
    sample_rows = [dict(zip(headers, r)) for r in rows[:SAMPLE_SIZE]]
    return ProfileResult(
        path=str(path), format="xlsx", rows=len(rows), cols=len(headers),
        columns=columns, sample_rows=sample_rows, truncated=truncated,
        limits=limits,
    )


def _profile_json(path: Path, max_rows: int) -> ProfileResult:
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    # Si dict, chercher une clé liste classique
    if isinstance(data, dict):
        for k in ("data", "results", "items", "features", "records"):
            if k in data and isinstance(data[k], list):
                data = data[k]
                break
    if not isinstance(data, list):
        return ProfileResult(
            path=str(path), format="json", rows=0, cols=0,
            columns=[], sample_rows=[], truncated=False,
            error="JSON non tabulaire (attendu : liste d'objets ou dict avec clé data/results/items)",
        )
    if not data:
        return ProfileResult(
            path=str(path), format="json", rows=0, cols=0,
            columns=[], sample_rows=[], truncated=False,
        )
    truncated = False
    limits: List[str] = []
    if len(data) > max_rows:
        data = data[:max_rows]
        truncated = True
        limits.append(f"Tronqué à {max_rows} lignes")
    # Union des clés sur les 100 premiers (nettoie BOM résiduel)
    headers: List[str] = []
    seen = set()
    for item in data[:100]:
        if isinstance(item, dict):
            for k in item:
                kn = _strip_bom(k).strip() if isinstance(k, str) else k
                if kn not in seen:
                    seen.add(kn)
                    headers.append(kn)
    rows: List[List[str]] = []
    for item in data:
        if isinstance(item, dict):
            # Re-map les clés nettoyées
            clean_item = {(_strip_bom(k).strip() if isinstance(k, str) else k): v for k, v in item.items()}
            rows.append([str(clean_item.get(h, "")) for h in headers])
    columns = _profile_columns(headers, rows)
    sample_rows = [dict(zip(headers, r)) for r in rows[:SAMPLE_SIZE]]
    return ProfileResult(
        path=str(path), format="json", rows=len(rows), cols=len(headers),
        columns=columns, sample_rows=sample_rows, truncated=truncated,
        limits=limits,
    )


# ─── Provenance sidecar ─────────────────────────────────────────────────

def _load_provenance(path: Path) -> Optional[dict]:
    """Lit `<file>.datagouv.json` si présent."""
    sidecar = path.with_suffix(path.suffix + ".datagouv.json")
    if not sidecar.exists():
        return None
    try:
        with sidecar.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


# ─── API publique ───────────────────────────────────────────────────────

# ─── V2.2 : filter/sort/limit (réutilise les readers V2.1) ──────────────

# Opérateurs whitelistés (anti-injection : pas d'eval/pd.query brut)
_ALLOWED_OPS: frozenset = frozenset({
    "==", "!=", ">", "<", ">=", "<=", "contains", "in", "startswith", "not_in",
})

# Bornes output
MAX_FILTER_LIMIT = 1000          # plafond dur, le LLM ne peut pas dépasser
DEFAULT_FILTER_LIMIT = 50
MAX_FILTER_OUTPUT_CHARS = 8000


@dataclass
class FilterResult:
    path: str
    format: str
    columns: List[str]
    rows: List[List[str]]            # lignes filtrées (déjà limitées)
    total_matched: int                # nb lignes matchées AVANT limit
    total_scanned: int                # nb lignes parcourues (peut être < total fichier si truncated)
    truncated_at_load: bool           # True si fichier > max_rows_load
    truncated_at_limit: bool          # True si total_matched > limit
    encoding_used: Optional[str] = None
    separator_used: Optional[str] = None
    error: Optional[str] = None


def _load_tabular(path: Path, max_rows: int):
    """Charge un fichier tabulaire en (headers, rows, encoding, sep).

    Réutilise les readers de profile_file.
    Retourne (None, None, None, None, error_msg) si échec.
    """
    if not path.exists():
        return None, None, None, None, f"Fichier introuvable : {path}"
    ext = path.suffix.lower().lstrip(".")
    if ext == "csv":
        encoding = _detect_encoding(path)
        separator = _detect_separator(path, encoding)
        with path.open("r", encoding=encoding, errors="replace", newline="") as f:
            reader = csv.reader(f, delimiter=separator)
            try:
                headers = _clean_headers(next(reader))
            except StopIteration:
                return [], [], encoding, separator, None
            rows: List[List[str]] = []
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                rows.append(row)
        return headers, rows, encoding, separator, None
    if ext in ("xlsx", "xlsm"):
        try:
            import openpyxl
        except ImportError:
            return None, None, None, None, "openpyxl non installé"
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            first = next(rows_iter)
        except StopIteration:
            wb.close()
            return [], [], None, None, None
        headers = _clean_headers([str(c) if c is not None else "" for c in first])
        rows: List[List[str]] = []
        for i, row in enumerate(rows_iter):
            if i >= max_rows:
                break
            rows.append([str(c) if c is not None else "" for c in row])
        wb.close()
        return headers, rows, None, None, None
    if ext == "json":
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            for k in ("data", "results", "items", "features", "records"):
                if k in data and isinstance(data[k], list):
                    data = data[k]
                    break
        if not isinstance(data, list):
            return None, None, None, None, "JSON non tabulaire"
        data = data[:max_rows]
        headers: List[str] = []
        seen = set()
        for item in data[:100]:
            if isinstance(item, dict):
                for k in item:
                    kn = _strip_bom(k).strip() if isinstance(k, str) else k
                    if kn not in seen:
                        seen.add(kn)
                        headers.append(kn)
        rows: List[List[str]] = []
        for item in data:
            if isinstance(item, dict):
                clean_item = {(_strip_bom(k).strip() if isinstance(k, str) else k): v for k, v in item.items()}
                rows.append([str(clean_item.get(h, "")) for h in headers])
        return headers, rows, None, None, None
    if ext == "xls":
        return None, None, None, None, "Format .xls (Excel legacy) non supporté"
    return None, None, None, None, f"Format non supporté : .{ext}"


def _coerce_value(v: str):
    """Tente de convertir une chaîne en int/float pour comparaison numérique.

    Retourne (number, is_numeric). Si non numérique → (string, False).
    Accepte la virgule décimale FR.
    """
    if v is None or v == "":
        return v, False
    s = v.strip()
    if _INT_RE.match(s):
        try:
            return int(s), True
        except ValueError:
            return s, False
    if _FLOAT_RE.match(s):
        try:
            return float(s.replace(",", ".")), True
        except ValueError:
            return s, False
    return s, False


def _apply_condition(cell: str, op: str, target) -> bool:
    """Évalue une condition filtre. Cell est toujours une string (origine CSV/XLSX/JSON).

    Pour les ops numériques (>, <, >=, <=), tente coercion des deux côtés.
    Pour ==/!=, compare en string par défaut, fallback numérique si target est num.
    """
    if op == "in":
        if not isinstance(target, (list, tuple, set)):
            target = [target]
        return cell in [str(t) for t in target]
    if op == "not_in":
        if not isinstance(target, (list, tuple, set)):
            target = [target]
        return cell not in [str(t) for t in target]
    if op == "contains":
        return str(target).lower() in (cell or "").lower()
    if op == "startswith":
        return (cell or "").lower().startswith(str(target).lower())

    # Comparateurs : tenter conversion numérique côté target
    target_str = str(target)
    if op in ("==", "!="):
        # Comparaison exacte string par défaut
        if op == "==":
            return cell == target_str
        return cell != target_str
    # Ops numériques
    if op in (">", "<", ">=", "<="):
        cell_v, cell_is_num = _coerce_value(cell)
        target_v, target_is_num = _coerce_value(target_str)
        if cell_is_num and target_is_num:
            a, b = cell_v, target_v
        else:
            # Fallback comparaison string (utile pour dates ISO triables lex)
            a, b = cell, target_str
        if op == ">":
            return a > b
        if op == "<":
            return a < b
        if op == ">=":
            return a >= b
        if op == "<=":
            return a <= b
    raise ValueError(f"Opérateur inconnu : {op}")


class FilterError(ValueError):
    """Erreur de filtre côté contrat (colonne inexistante, opérateur interdit, etc)."""


def _validate_conditions(headers: List[str], conditions: list) -> None:
    """Valide la structure des conditions. Lève FilterError si invalide."""
    headers_set = set(headers)
    for i, cond in enumerate(conditions):
        if not isinstance(cond, dict):
            raise FilterError(f"Condition #{i+1} doit être un objet (col/op/value)")
        col = cond.get("col") or cond.get("column")
        op = cond.get("op")
        if not col:
            raise FilterError(f"Condition #{i+1} : `col` manquant")
        if col not in headers_set:
            # suggestion fuzzy simple (préfixe ou substring)
            close = [h for h in headers if col.lower() in h.lower() or h.lower() in col.lower()]
            hint = f" Suggestions : {close[:3]}" if close else f" Colonnes disponibles : {headers[:10]}"
            raise FilterError(f"Colonne inconnue : `{col}`.{hint}")
        if op not in _ALLOWED_OPS:
            raise FilterError(
                f"Opérateur `{op}` interdit. Whitelist : {sorted(_ALLOWED_OPS)}"
            )
        if "value" not in cond:
            raise FilterError(f"Condition #{i+1} ({col} {op}) : `value` manquant")


def _validate_sort(headers: List[str], sort_specs: list) -> list:
    """Normalise les specs de tri. Retourne [(col, reverse_bool), ...]."""
    out = []
    headers_set = set(headers)
    for s in sort_specs:
        if isinstance(s, str):
            # "-col" = desc, "col" = asc, "+col" = asc
            if s.startswith("-"):
                col, reverse = s[1:], True
            elif s.startswith("+"):
                col, reverse = s[1:], False
            else:
                col, reverse = s, False
        elif isinstance(s, dict):
            col = s.get("col") or s.get("column")
            order = (s.get("order") or "asc").lower()
            reverse = order in ("desc", "descending")
        else:
            raise FilterError(f"Spec de tri invalide : {s!r}")
        if not col:
            raise FilterError(f"Spec de tri : `col` manquant ({s!r})")
        if col not in headers_set:
            raise FilterError(
                f"Colonne de tri inconnue : `{col}`. Colonnes : {headers[:10]}"
            )
        out.append((col, reverse))
    return out


def filter_rows(
    path: Path,
    *,
    where: Optional[list] = None,
    sort: Optional[list] = None,
    limit: int = DEFAULT_FILTER_LIMIT,
    max_rows_load: Optional[int] = None,
) -> FilterResult:
    """Filtre + trie + limite un fichier tabulaire.

    Args:
        path: chemin du fichier
        where: liste de conditions {"col": str, "op": str, "value": any}
        sort: liste de specs ("-col" / "+col" / {"col": str, "order": "asc|desc"})
        limit: max lignes retournées (borné à MAX_FILTER_LIMIT)
        max_rows_load: max lignes chargées (défaut env LUMENA_DATA_WB_MAX_ROWS)
    """
    path = Path(path)
    max_load = max_rows_load if max_rows_load is not None else _max_rows_default()
    limit = max(1, min(int(limit or DEFAULT_FILTER_LIMIT), MAX_FILTER_LIMIT))

    headers, rows, enc, sep, err = _load_tabular(path, max_load)
    if err:
        return FilterResult(
            path=str(path), format=path.suffix.lower().lstrip("."),
            columns=[], rows=[], total_matched=0, total_scanned=0,
            truncated_at_load=False, truncated_at_limit=False,
            error=err,
        )
    if not headers:
        return FilterResult(
            path=str(path), format=path.suffix.lower().lstrip("."),
            columns=[], rows=[], total_matched=0, total_scanned=0,
            truncated_at_load=False, truncated_at_limit=False,
            encoding_used=enc, separator_used=sep,
        )

    where = where or []
    sort_specs = sort or []
    _validate_conditions(headers, where)
    sort_normalized = _validate_sort(headers, sort_specs)

    col_index = {h: i for i, h in enumerate(headers)}

    # Filtre
    def row_passes(row: List[str]) -> bool:
        for cond in where:
            col = cond.get("col") or cond.get("column")
            op = cond["op"]
            val = cond["value"]
            idx = col_index[col]
            cell = row[idx] if idx < len(row) else ""
            if not _apply_condition(cell, op, val):
                return False
        return True

    matched = [r for r in rows if row_passes(r)]
    total_matched = len(matched)

    # Tri
    if sort_normalized:
        for col, reverse in reversed(sort_normalized):
            idx = col_index[col]
            def _key(r, _i=idx):
                cell = r[_i] if _i < len(r) else ""
                v, is_num = _coerce_value(cell)
                # Tri stable mixte : numériques d'abord (par valeur), puis textes
                return (0 if is_num else 1, v if is_num else (cell or ""))
            matched.sort(key=_key, reverse=reverse)

    truncated_at_limit = total_matched > limit
    matched = matched[:limit]

    return FilterResult(
        path=str(path), format=path.suffix.lower().lstrip("."),
        columns=headers, rows=matched,
        total_matched=total_matched, total_scanned=len(rows),
        truncated_at_load=len(rows) >= max_load,
        truncated_at_limit=truncated_at_limit,
        encoding_used=enc, separator_used=sep,
    )


# ─── V2.3 : aggregate / unique_values ───────────────────────────────────

_ALLOWED_AGGS: frozenset = frozenset({
    "count", "sum", "mean", "avg", "average", "min", "max", "median",
})

MAX_GROUPS = 1000     # plafond dur sur le nb de groupes retournés
DEFAULT_GROUPS = 100
MAX_UNIQUE_LIMIT = 1000
DEFAULT_UNIQUE_LIMIT = 100


@dataclass
class AggregateResult:
    path: str
    format: str
    group_by: List[str]
    agg: str
    agg_col: Optional[str]
    rows: List[dict]          # [{group_col1: v, ..., result: x, _count: n}, ...]
    total_groups: int
    total_scanned: int
    truncated_at_load: bool
    truncated_at_limit: bool
    encoding_used: Optional[str] = None
    separator_used: Optional[str] = None
    error: Optional[str] = None


@dataclass
class UniqueValuesResult:
    path: str
    format: str
    column: str
    values: List[tuple]       # [(value, count), ...]
    total_unique: int
    total_scanned: int
    truncated_at_load: bool
    truncated_at_limit: bool
    encoding_used: Optional[str] = None
    separator_used: Optional[str] = None
    error: Optional[str] = None


def _normalize_agg(agg: str) -> str:
    a = (agg or "").lower().strip()
    if a in ("avg", "average"):
        return "mean"
    return a


def _reduce(values: List[float], agg: str) -> Optional[float]:
    if not values:
        return None
    if agg == "count":
        return float(len(values))
    if agg == "sum":
        return sum(values)
    if agg == "mean":
        return sum(values) / len(values)
    if agg == "min":
        return min(values)
    if agg == "max":
        return max(values)
    if agg == "median":
        sv = sorted(values)
        n = len(sv)
        mid = n // 2
        return sv[mid] if n % 2 else (sv[mid - 1] + sv[mid]) / 2.0
    raise ValueError(f"Agg inconnue : {agg}")


def aggregate_data(
    path: Path,
    *,
    group_by: List[str],
    agg: str,
    agg_col: Optional[str] = None,
    where: Optional[list] = None,
    sort: Optional[list] = None,
    limit: int = DEFAULT_GROUPS,
    max_rows_load: Optional[int] = None,
) -> AggregateResult:
    """Groupby + agrégation simple.

    - `group_by` : 1 à 3 colonnes
    - `agg` ∈ {count, sum, mean, min, max, median} (avg/average = mean)
    - `agg_col` requis sauf pour count
    - `where` (optionnel) : pré-filtre identique à filter_rows
    - `sort` (optionnel) : tri sur les colonnes du résultat (`group_by` ou `result`)
    - `limit` borné à MAX_GROUPS
    """
    path = Path(path)
    max_load = max_rows_load if max_rows_load is not None else _max_rows_default()
    limit = max(1, min(int(limit or DEFAULT_GROUPS), MAX_GROUPS))
    agg_norm = _normalize_agg(agg)
    if agg_norm not in _ALLOWED_AGGS - {"avg", "average"}:
        raise FilterError(
            f"Agrégation `{agg}` interdite. Whitelist : "
            f"{sorted({'count', 'sum', 'mean', 'min', 'max', 'median'})}"
        )
    if agg_norm != "count" and not agg_col:
        raise FilterError(
            f"Agrégation `{agg_norm}` nécessite `agg_col` (colonne numérique)."
        )

    if isinstance(group_by, str):
        group_by = [group_by]
    if not group_by:
        raise FilterError("`group_by` ne peut pas être vide (V2.3).")
    if len(group_by) > 3:
        raise FilterError("`group_by` limité à 3 colonnes en V2.3.")

    headers, rows, enc, sep, err = _load_tabular(path, max_load)
    if err:
        return AggregateResult(
            path=str(path), format=path.suffix.lower().lstrip("."),
            group_by=group_by, agg=agg_norm, agg_col=agg_col,
            rows=[], total_groups=0, total_scanned=0,
            truncated_at_load=False, truncated_at_limit=False,
            error=err,
        )
    if not headers:
        return AggregateResult(
            path=str(path), format=path.suffix.lower().lstrip("."),
            group_by=group_by, agg=agg_norm, agg_col=agg_col,
            rows=[], total_groups=0, total_scanned=0,
            truncated_at_load=False, truncated_at_limit=False,
            encoding_used=enc, separator_used=sep,
        )

    # Validation colonnes
    headers_set = set(headers)
    for col in group_by:
        if col not in headers_set:
            close = [h for h in headers if col.lower() in h.lower() or h.lower() in col.lower()]
            hint = f" Suggestions : {close[:3]}" if close else f" Colonnes : {headers[:10]}"
            raise FilterError(f"Colonne de groupage inconnue : `{col}`.{hint}")
    if agg_col and agg_col not in headers_set:
        close = [h for h in headers if agg_col.lower() in h.lower() or h.lower() in agg_col.lower()]
        hint = f" Suggestions : {close[:3]}" if close else f" Colonnes : {headers[:10]}"
        raise FilterError(f"Colonne d'agrégation inconnue : `{agg_col}`.{hint}")

    where = where or []
    _validate_conditions(headers, where)
    col_index = {h: i for i, h in enumerate(headers)}

    # Pré-filtre
    def passes(row):
        for cond in where:
            col = cond.get("col") or cond.get("column")
            op = cond["op"]
            val = cond["value"]
            idx = col_index[col]
            cell = row[idx] if idx < len(row) else ""
            if not _apply_condition(cell, op, val):
                return False
        return True

    # Groupage
    gb_idx = [col_index[c] for c in group_by]
    agg_idx = col_index[agg_col] if agg_col else None
    groups: dict = {}  # tuple_key → list of values (ou count)

    for row in rows:
        if where and not passes(row):
            continue
        key = tuple(row[i] if i < len(row) else "" for i in gb_idx)
        if agg_norm == "count":
            groups[key] = groups.get(key, 0) + 1
        else:
            cell = row[agg_idx] if agg_idx is not None and agg_idx < len(row) else ""
            v, is_num = _coerce_value(cell)
            if is_num:
                groups.setdefault(key, []).append(float(v))

    # Reduce
    out_rows: List[dict] = []
    for key, payload in groups.items():
        row_dict = dict(zip(group_by, key))
        if agg_norm == "count":
            row_dict["result"] = float(payload)
            row_dict["_count"] = int(payload)
        else:
            row_dict["result"] = _reduce(payload, agg_norm)
            row_dict["_count"] = len(payload)
        out_rows.append(row_dict)

    total_groups = len(out_rows)

    # Sort post-agrégation
    if sort:
        sort_specs = sort if isinstance(sort, list) else [sort]
        for spec in reversed(sort_specs):
            if isinstance(spec, str):
                if spec.startswith("-"):
                    col, reverse = spec[1:], True
                elif spec.startswith("+"):
                    col, reverse = spec[1:], False
                else:
                    col, reverse = spec, False
            elif isinstance(spec, dict):
                col = spec.get("col") or spec.get("column")
                reverse = (spec.get("order") or "asc").lower() in ("desc", "descending")
            else:
                raise FilterError(f"Spec de tri invalide : {spec!r}")
            if col not in (set(group_by) | {"result", "_count"}):
                raise FilterError(
                    f"Tri impossible sur `{col}` : choisir parmi {group_by + ['result', '_count']}."
                )
            out_rows.sort(
                key=lambda r, _c=col: (r.get(_c) is None, r.get(_c) or 0 if _c in ("result", "_count") else (r.get(_c) or "")),
                reverse=reverse,
            )

    truncated_at_limit = total_groups > limit
    out_rows = out_rows[:limit]

    return AggregateResult(
        path=str(path), format=path.suffix.lower().lstrip("."),
        group_by=group_by, agg=agg_norm, agg_col=agg_col,
        rows=out_rows, total_groups=total_groups, total_scanned=len(rows),
        truncated_at_load=len(rows) >= max_load,
        truncated_at_limit=truncated_at_limit,
        encoding_used=enc, separator_used=sep,
    )


def unique_values(
    path: Path,
    *,
    column: str,
    limit: int = DEFAULT_UNIQUE_LIMIT,
    max_rows_load: Optional[int] = None,
    include_empty: bool = False,
) -> UniqueValuesResult:
    """Liste les valeurs distinctes d'une colonne avec leur fréquence.

    Tri : fréquence décroissante.
    """
    path = Path(path)
    max_load = max_rows_load if max_rows_load is not None else _max_rows_default()
    limit = max(1, min(int(limit or DEFAULT_UNIQUE_LIMIT), MAX_UNIQUE_LIMIT))

    headers, rows, enc, sep, err = _load_tabular(path, max_load)
    if err:
        return UniqueValuesResult(
            path=str(path), format=path.suffix.lower().lstrip("."),
            column=column, values=[], total_unique=0, total_scanned=0,
            truncated_at_load=False, truncated_at_limit=False,
            error=err,
        )
    if not headers:
        return UniqueValuesResult(
            path=str(path), format=path.suffix.lower().lstrip("."),
            column=column, values=[], total_unique=0, total_scanned=0,
            truncated_at_load=False, truncated_at_limit=False,
            encoding_used=enc, separator_used=sep,
        )

    if column not in headers:
        close = [h for h in headers if column.lower() in h.lower() or h.lower() in column.lower()]
        hint = f" Suggestions : {close[:3]}" if close else f" Colonnes : {headers[:10]}"
        raise FilterError(f"Colonne inconnue : `{column}`.{hint}")

    idx = headers.index(column)
    counts: dict = {}
    for row in rows:
        v = row[idx] if idx < len(row) else ""
        if not include_empty and (v is None or v == ""):
            continue
        counts[v] = counts.get(v, 0) + 1

    sorted_pairs = sorted(counts.items(), key=lambda kv: (-kv[1], kv[0]))
    total_unique = len(sorted_pairs)
    truncated = total_unique > limit
    sorted_pairs = sorted_pairs[:limit]

    return UniqueValuesResult(
        path=str(path), format=path.suffix.lower().lstrip("."),
        column=column, values=sorted_pairs,
        total_unique=total_unique, total_scanned=len(rows),
        truncated_at_load=len(rows) >= max_load,
        truncated_at_limit=truncated,
        encoding_used=enc, separator_used=sep,
    )


# ─── V2.4 : export transformé ───────────────────────────────────────────

_ALLOWED_EXPORT_FORMATS: frozenset = frozenset({"csv", "json", "xlsx"})
MAX_EXPORT_ROWS = 100_000  # plafond dur sortie fichier


@dataclass
class ExportResult:
    source_path: str
    output_path: str
    sidecar_path: str
    output_format: str
    rows_exported: int
    operations: dict  # {where, group_by, agg, agg_col, sort, limit, columns}
    truncated_at_load: bool
    truncated_at_export: bool
    preview_headers: List[str] = field(default_factory=list)
    preview_rows: List[List[str]] = field(default_factory=list)  # 5 premières lignes
    error: Optional[str] = None


def _project_columns(headers: List[str], rows: List[List[str]], columns: Optional[List[str]]):
    """Si `columns` est fourni, ne garde que ces colonnes dans l'ordre demandé."""
    if not columns:
        return headers, rows
    headers_set = set(headers)
    missing = [c for c in columns if c not in headers_set]
    if missing:
        raise FilterError(
            f"Colonne(s) inconnue(s) pour projection : {missing}. Colonnes : {headers[:10]}"
        )
    col_index = {h: i for i, h in enumerate(headers)}
    indices = [col_index[c] for c in columns]
    new_rows = [
        [row[i] if i < len(row) else "" for i in indices] for row in rows
    ]
    return list(columns), new_rows


def _write_csv(path: Path, headers: List[str], rows: List[List[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=",")
        writer.writerow(headers)
        for r in rows:
            writer.writerow(r)


def _write_json(path: Path, headers: List[str], rows: List[List[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    data = [dict(zip(headers, r)) for r in rows]
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _write_xlsx(path: Path, headers: List[str], rows: List[List[str]]) -> None:
    import openpyxl
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(str(path))


def export_data(
    source: Path,
    output_path: Path,
    *,
    output_format: str,
    where: Optional[list] = None,
    group_by: Optional[list] = None,
    agg: Optional[str] = None,
    agg_col: Optional[str] = None,
    sort: Optional[list] = None,
    limit: Optional[int] = None,
    columns: Optional[List[str]] = None,
    max_rows_load: Optional[int] = None,
) -> ExportResult:
    """Filtre/agrège un fichier source puis exporte le résultat.

    Sortie : CSV / JSON / XLSX. Sidecar `<file>.export_meta.json` avec
    operations + provenance.
    """
    source = Path(source)
    output_path = Path(output_path)
    fmt = (output_format or "").lower().strip().lstrip(".")
    if fmt not in _ALLOWED_EXPORT_FORMATS:
        raise FilterError(
            f"Format de sortie `{output_format}` interdit. "
            f"Whitelist : {sorted(_ALLOWED_EXPORT_FORMATS)}"
        )

    operations = {
        "where": where or [],
        "group_by": list(group_by) if group_by else [],
        "agg": agg,
        "agg_col": agg_col,
        "sort": list(sort) if sort else [],
        "limit": limit,
        "columns": list(columns) if columns else [],
        "output_format": fmt,
    }

    # Pipeline : agrégation OU filter+project
    if group_by:
        if not agg:
            raise FilterError("`agg` requis quand `group_by` est fourni.")
        agg_result = aggregate_data(
            source,
            group_by=group_by, agg=agg, agg_col=agg_col,
            where=where, sort=sort,
            limit=limit if limit is not None else MAX_EXPORT_ROWS,
            max_rows_load=max_rows_load,
        )
        if agg_result.error:
            return ExportResult(
                source_path=str(source), output_path=str(output_path), sidecar_path="",
                output_format=fmt, rows_exported=0, operations=operations,
                truncated_at_load=False, truncated_at_export=False,
                error=agg_result.error,
            )
        export_headers = list(agg_result.group_by) + ["result", "_count"]
        export_rows = [
            [str(row.get(c, "")) for c in export_headers]
            for row in agg_result.rows
        ]
        rows_exported = len(export_rows)
        truncated_load = agg_result.truncated_at_load
        truncated_export = agg_result.truncated_at_limit
    else:
        filter_result = filter_rows(
            source,
            where=where, sort=sort,
            limit=limit if limit is not None else MAX_EXPORT_ROWS,
            max_rows_load=max_rows_load,
        )
        if filter_result.error:
            return ExportResult(
                source_path=str(source), output_path=str(output_path), sidecar_path="",
                output_format=fmt, rows_exported=0, operations=operations,
                truncated_at_load=False, truncated_at_export=False,
                error=filter_result.error,
            )
        export_headers, export_rows = _project_columns(
            filter_result.columns, filter_result.rows, columns
        )
        rows_exported = len(export_rows)
        truncated_load = filter_result.truncated_at_load
        truncated_export = filter_result.truncated_at_limit

    # Garde-fou taille fichier de sortie
    if rows_exported > MAX_EXPORT_ROWS:
        export_rows = export_rows[:MAX_EXPORT_ROWS]
        rows_exported = MAX_EXPORT_ROWS
        truncated_export = True

    # Écriture
    if fmt == "csv":
        _write_csv(output_path, export_headers, export_rows)
    elif fmt == "json":
        _write_json(output_path, export_headers, export_rows)
    elif fmt == "xlsx":
        _write_xlsx(output_path, export_headers, export_rows)

    # Hash source
    import hashlib as _hashlib
    h = _hashlib.md5()
    with source.open("rb") as f:
        for chunk in iter(lambda: f.read(64 * 1024), b""):
            h.update(chunk)
    source_md5 = h.hexdigest()

    # Sidecar provenance
    from datetime import datetime, timezone
    sidecar = output_path.with_suffix(output_path.suffix + ".export_meta.json")
    meta = {
        "schema_version": 1,
        "source_path": str(source),
        "source_md5": source_md5,
        "output_path": str(output_path),
        "output_format": fmt,
        "rows_exported": rows_exported,
        "headers": export_headers,
        "operations": operations,
        "truncated_at_load": truncated_load,
        "truncated_at_export": truncated_export,
        "exported_at": datetime.now(timezone.utc).isoformat(),
    }
    sidecar.write_text(
        json.dumps(meta, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    return ExportResult(
        source_path=str(source),
        output_path=str(output_path),
        sidecar_path=str(sidecar),
        output_format=fmt,
        rows_exported=rows_exported,
        operations=operations,
        truncated_at_load=truncated_load,
        truncated_at_export=truncated_export,
        preview_headers=list(export_headers),
        preview_rows=[list(r) for r in export_rows[:5]],
    )


# ─── V3.4 : data_join ───────────────────────────────────────────────────

_ALLOWED_JOIN_TYPES: frozenset = frozenset({"inner", "left", "right", "outer"})
MAX_JOIN_OUTPUT_ROWS = 100_000  # plafond dur sur lignes jointes
DEFAULT_JOIN_LIMIT = 1000       # défaut visible (output borné)


@dataclass
class JoinResult:
    left_path: str
    right_path: str
    how: str
    on_left: str
    on_right: str
    columns: List[str]
    rows: List[List[str]]
    total_left: int
    total_right: int
    total_joined: int      # avant truncation limit
    truncated_at_left: bool
    truncated_at_right: bool
    truncated_at_output: bool
    error: Optional[str] = None


def _dedupe_columns(left_cols: List[str], right_cols: List[str], join_keys: tuple):
    """Construit la liste finale de colonnes en suffixant les collisions.

    - left_cols restent comme tels
    - right_cols : si nom déjà pris, suffixe `_right` (sauf pour la clé de join)
    Retourne (final_cols, left_indices, right_indices_with_keys) :
      - final_cols : liste ordonnée des colonnes du résultat
      - left_map : index → position dans final
      - right_map : index → position dans final (None si exclu)
    """
    final: List[str] = list(left_cols)
    seen = set(left_cols)
    right_positions = []
    on_left, on_right = join_keys
    for c in right_cols:
        if c == on_right and on_right == on_left:
            # même clé de jointure → ne pas dupliquer
            right_positions.append(None)
            continue
        if c in seen:
            new_name = f"{c}_right"
            # éviter collisions à répétition
            suffix = 2
            while new_name in seen:
                new_name = f"{c}_right{suffix}"
                suffix += 1
            final.append(new_name)
            right_positions.append(len(final) - 1)
            seen.add(new_name)
        else:
            final.append(c)
            right_positions.append(len(final) - 1)
            seen.add(c)
    return final, right_positions


def data_join(
    left_path: Path,
    right_path: Path,
    *,
    on_left: str,
    on_right: Optional[str] = None,
    how: str = "inner",
    limit: int = DEFAULT_JOIN_LIMIT,
    max_rows_load: Optional[int] = None,
) -> JoinResult:
    """Jointure simple entre deux fichiers tabulaires existants.

    Args:
        left_path / right_path : chemins absolus des fichiers.
        on_left : colonne de jointure dans le fichier gauche.
        on_right : colonne dans le fichier droit (défaut = on_left).
        how : 'inner' | 'left' | 'right' | 'outer'.
        limit : lignes max retournées (cap dur MAX_JOIN_OUTPUT_ROWS).
        max_rows_load : lignes max chargées par fichier.
    """
    left_path = Path(left_path)
    right_path = Path(right_path)
    on_right = on_right or on_left
    how_norm = (how or "inner").lower().strip()
    if how_norm not in _ALLOWED_JOIN_TYPES:
        raise FilterError(
            f"Type de jointure `{how}` interdit. "
            f"Whitelist : {sorted(_ALLOWED_JOIN_TYPES)}"
        )
    max_load = max_rows_load if max_rows_load is not None else _max_rows_default()
    limit = max(1, min(int(limit or DEFAULT_JOIN_LIMIT), MAX_JOIN_OUTPUT_ROWS))

    # Load left
    left_headers, left_rows, _, _, err = _load_tabular(left_path, max_load)
    if err:
        return JoinResult(
            left_path=str(left_path), right_path=str(right_path),
            how=how_norm, on_left=on_left, on_right=on_right,
            columns=[], rows=[], total_left=0, total_right=0, total_joined=0,
            truncated_at_left=False, truncated_at_right=False,
            truncated_at_output=False,
            error=f"left: {err}",
        )
    # Load right
    right_headers, right_rows, _, _, err = _load_tabular(right_path, max_load)
    if err:
        return JoinResult(
            left_path=str(left_path), right_path=str(right_path),
            how=how_norm, on_left=on_left, on_right=on_right,
            columns=[], rows=[], total_left=len(left_rows or []), total_right=0,
            total_joined=0,
            truncated_at_left=False, truncated_at_right=False,
            truncated_at_output=False,
            error=f"right: {err}",
        )

    # Validation des clés
    if on_left not in (left_headers or []):
        close = [h for h in (left_headers or []) if on_left.lower() in h.lower()]
        hint = f" Suggestions : {close[:3]}" if close else f" Colonnes left : {left_headers[:10]}"
        raise FilterError(f"Colonne `on_left={on_left}` absente du fichier gauche.{hint}")
    if on_right not in (right_headers or []):
        close = [h for h in (right_headers or []) if on_right.lower() in h.lower()]
        hint = f" Suggestions : {close[:3]}" if close else f" Colonnes right : {right_headers[:10]}"
        raise FilterError(f"Colonne `on_right={on_right}` absente du fichier droit.{hint}")

    left_idx_join = left_headers.index(on_left)
    right_idx_join = right_headers.index(on_right)

    # Index droit (clé → liste de lignes)
    right_index: dict = {}
    for r in right_rows:
        key = r[right_idx_join] if right_idx_join < len(r) else ""
        right_index.setdefault(key, []).append(r)

    final_cols, right_positions = _dedupe_columns(
        left_headers, right_headers, (on_left, on_right)
    )
    total_cols = len(final_cols)

    def _build_row(left_row, right_row):
        out = [""] * total_cols
        # left
        for i, c in enumerate(left_row[: len(left_headers)]):
            out[i] = c if c is not None else ""
        # right (sauf colonnes None = clé de join exclue)
        if right_row is not None:
            for j, pos in enumerate(right_positions):
                if pos is None:
                    continue
                if j < len(right_row):
                    out[pos] = right_row[j] if right_row[j] is not None else ""
        return out

    # Construction selon how
    joined: List[List[str]] = []
    matched_right_keys: set = set()
    for lrow in left_rows:
        key = lrow[left_idx_join] if left_idx_join < len(lrow) else ""
        right_matches = right_index.get(key)
        if right_matches:
            matched_right_keys.add(key)
            for rrow in right_matches:
                joined.append(_build_row(lrow, rrow))
                if len(joined) >= MAX_JOIN_OUTPUT_ROWS:
                    break
        else:
            if how_norm in ("left", "outer"):
                joined.append(_build_row(lrow, None))
        if len(joined) >= MAX_JOIN_OUTPUT_ROWS:
            break

    # right / outer : lignes droite non matchées
    if how_norm in ("right", "outer") and len(joined) < MAX_JOIN_OUTPUT_ROWS:
        for rrow in right_rows:
            key = rrow[right_idx_join] if right_idx_join < len(rrow) else ""
            if key in matched_right_keys and how_norm == "right":
                # déjà émis quand on a parcouru left (sauf si left ne matchait pas)
                # Pour right pur : on veut TOUTES les lignes droite ; il faut donc tout reconstruire.
                # Pour simplicité : right pur émet toutes les lignes droite, avec left vide si pas matched.
                pass
            if how_norm == "right":
                # toutes les lignes droite
                left_matches = [
                    lr for lr in left_rows
                    if (lr[left_idx_join] if left_idx_join < len(lr) else "") == key
                ]
                if left_matches:
                    # déjà émises au-dessus (left iter)
                    if key in matched_right_keys:
                        continue
                joined.append(_build_row([""] * len(left_headers), rrow))
            elif how_norm == "outer" and key not in matched_right_keys:
                joined.append(_build_row([""] * len(left_headers), rrow))
            if len(joined) >= MAX_JOIN_OUTPUT_ROWS:
                break

    total_joined = len(joined)
    truncated_output = total_joined > limit
    joined = joined[:limit]

    return JoinResult(
        left_path=str(left_path), right_path=str(right_path),
        how=how_norm, on_left=on_left, on_right=on_right,
        columns=final_cols, rows=joined,
        total_left=len(left_rows), total_right=len(right_rows),
        total_joined=total_joined,
        truncated_at_left=len(left_rows) >= max_load,
        truncated_at_right=len(right_rows) >= max_load,
        truncated_at_output=truncated_output,
    )


def profile_file(path: Path, max_rows: Optional[int] = None) -> ProfileResult:
    """Profile un fichier tabulaire (CSV/XLSX/JSON)."""
    path = Path(path)
    if not path.exists():
        return ProfileResult(
            path=str(path), format="?", rows=0, cols=0,
            columns=[], sample_rows=[], truncated=False,
            error=f"Fichier introuvable : {path}",
        )
    max_r = max_rows if max_rows is not None else _max_rows_default()
    ext = path.suffix.lower().lstrip(".")
    if ext == "csv":
        result = _profile_csv(path, max_r)
    elif ext in ("xlsx", "xlsm"):
        result = _profile_xlsx(path, max_r)
    elif ext == "json":
        result = _profile_json(path, max_r)
    elif ext == "xls":
        return ProfileResult(
            path=str(path), format=ext, rows=0, cols=0,
            columns=[], sample_rows=[], truncated=False,
            error=(
                "Format .xls (Excel legacy 97-2003) non supporté en V2.1. "
                "Réessayer avec un fichier .xlsx ou demander une autre resource au dataset."
            ),
        )
    else:
        return ProfileResult(
            path=str(path), format=ext or "?", rows=0, cols=0,
            columns=[], sample_rows=[], truncated=False,
            error=f"Format non supporté en V2.1 : .{ext} (supportés : csv, xlsx, json)",
        )
    result.provenance = _load_provenance(path)
    return result
